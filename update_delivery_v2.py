"""
OPENDUCK V3 - Delivery Status Update v2
Based on Amazon order history shared 2026-01-14
Works with MASTER_TRACKER sheet
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

TRACKER_PATH = r"C:\Users\matte\Desktop\Desktop OLD\AI\Università AI\courses\personal_project\robot_jarvis\OPENDUCK_V3_FINAL_TRACKER.xlsx"

# Status colors
COLORS = {
    'RICEVUTO': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'IN ARRIVO': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    'DA ORDINARE': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    'RESO': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
    'RIMOSSO': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
}

# Delivery status based on Amazon order history (2026-01-14)
# Format: keyword -> (status, delivery_date/note)
DELIVERY_STATUS = {
    # RICEVUTO (Delivered)
    'QIDI': ('RICEVUTO', 'Dic 2025'),
    'X-Max': ('RICEVUTO', 'Dic 2025'),
    'MG90S': ('RICEVUTO', '13/01'),
    'Raspberry Pi 4': ('RICEVUTO', '14/01'),
    'PCA9685': ('IN ARRIVO', '15/01'),
    'USB-C': ('IN ARRIVO', '15/01'),
    'Alluminio': ('IN ARRIVO', '15/01'),
    'Interruttore': ('RICEVUTO', '13/01'),
    'Porta batteria': ('RICEVUTO', '13/01'),
    'CABLEPELADO': ('RICEVUTO', '13/01'),
    'batteria 18650': ('RICEVUTO', '13/01'),
    'Limit Switch': ('RICEVUTO', '13/01'),
    'KW11': ('RICEVUTO', '13/01'),
    'Taiss': ('RICEVUTO', '13/01'),
    'WS2812': ('RICEVUTO', '14/01'),
    'NeoPixel': ('RICEVUTO', '14/01'),
    'XT30': ('RICEVUTO', '13-14/01'),
    'YIXISI': ('RICEVUTO', '14/01'),
    'Yiqigou': ('RICEVUTO', '13/01'),
    'HC-SR04': ('RICEVUTO', '13/01'),
    'Aihasd': ('RICEVUTO', '13/01'),
    'Ultrasonic': ('RICEVUTO', '13/01'),
    'Silicone': ('RICEVUTO', '13/01'),
    'Gruiqrd': ('RICEVUTO', '13/01'),
    'eSUN': ('RICEVUTO', '13/01'),
    'MAX98357': ('RICEVUTO', '13/01'),
    'AZDelivery MAX': ('RICEVUTO', '13/01'),
    'UBEC': ('RICEVUTO', '13/01'),
    'ZHITING': ('RICEVUTO', '13/01'),
    'ELEGOO': ('RICEVUTO', '14/01'),
    'Jumper': ('RICEVUTO', '14/01'),
    'Polymaker': ('RICEVUTO', '14/01'),
    'YINETTECH': ('RICEVUTO', '14/01'),
    'Servo Horn': ('RICEVUTO', '14/01'),
    'Servo Braccio': ('RICEVUTO', '14/01'),
    'saldatore': ('RICEVUTO', '14/01'),
    'Saldatore': ('RICEVUTO', '14/01'),
    'Viti Cilindriche': ('RICEVUTO', '14/01'),
    'SUNLU': ('RICEVUTO', '14/01'),
    'Silk': ('RICEVUTO', '14/01'),
    'BMS': ('RICEVUTO', '13/01'),
    '2S BMS': ('RICEVUTO', '13/01'),
    'TECNOIOT': ('RICEVUTO', '13/01'),
    'Kapton': ('RICEVUTO', '13/01'),
    'YUVKIN': ('RICEVUTO', '13/01'),
    'TPU': ('RICEVUTO', '13/01'),
    'JAYO': ('RICEVUTO', '13/01'),
    'HUAZIZ': ('RICEVUTO', '13/01'),
    'Servo Extension': ('RICEVUTO', '13/01'),
    'Isopropanol': ('RICEVUTO', '14/01'),
    'EQM': ('RICEVUTO', '14/01'),
    'Cuscinetti': ('RICEVUTO', '14/01'),
    'MR63ZZ': ('RICEVUTO', '14/01'),
    'Prusament': ('RICEVUTO', '14/01'),
    'Galaxy': ('RICEVUTO', '14/01'),
    'Enerpower': ('RICEVUTO', '14/01'),
    'Charger': ('RICEVUTO', '14/01'),
    'Li-ion Charger': ('RICEVUTO', '14/01'),

    # IN ARRIVO (In Transit)
    'Dophee': ('IN ARRIVO', '16/01'),
    'Glass Dome': ('IN ARRIVO', '16/01'),
    'TXS0108': ('IN ARRIVO', '22/01'),
    'Level Shifter': ('IN ARRIVO', '22/01'),
    'SanDisk': ('IN ARRIVO', '20/01'),
    'microSD': ('IN ARRIVO', '20/01'),
    'Paradisetronic': ('IN ARRIVO', '20/01'),
    'Speaker': ('IN ARRIVO', '20/01'),
    '2W 8': ('IN ARRIVO', '20/01'),
    'Altoparlante': ('IN ARRIVO', '20/01'),
    'BNO085': ('IN ARRIVO', '19-22/01'),
    'Adafruit': ('IN ARRIVO', '19-22/01'),
    'IMU': ('IN ARRIVO', '19-22/01'),
    '9-DOF': ('IN ARRIVO', '19-22/01'),
    'INMP441': ('IN ARRIVO', '15/01'),
    'AYWHP': ('IN ARRIVO', '15/01'),
    'I2S Microphone': ('IN ARRIVO', '15/01'),
    'ruthex': ('IN ARRIVO', '23/01'),
    'Heat Set': ('IN ARRIVO', '23/01'),
    'Flux Pen': ('IN ARRIVO', '6-12/02'),
    'FILO STAGNO': ('IN ARRIVO', '20/01'),
    'ETOPARS': ('IN ARRIVO', '15/01'),
    'Guaina': ('IN ARRIVO', '15/01'),

    # RESO (Returned)
    'Pi Zero': ('RESO', 'Reso'),
    'Zero 2W': ('RESO', 'Reso'),
    'Micro USB': ('RESO', 'Reso'),
    'PowerFast': ('RESO', 'Reso'),
    'AZDelivery Pi Zero': ('RESO', 'Reso'),
    'Camera Cable': ('RESO', 'Reso'),
}

# External orders (not Amazon)
EXTERNAL_STATUS = {
    'Molicel': ('DA ORDINARE', 'Vape Shop'),
    'P30B': ('DA ORDINARE', 'Vape Shop'),
    'INR18650': ('DA ORDINARE', 'Vape Shop'),
    'AI Camera': ('DA ORDINARE', 'Pimoroni'),
    'IMX500': ('DA ORDINARE', 'Pimoroni'),
    'FE-URT': ('DA ORDINARE', 'AliExpress'),
    'USB-UART': ('DA ORDINARE', 'AliExpress'),
    'STS3215': ('DA ORDINARE', 'Eckstein'),
    'Feetech': ('DA ORDINARE', 'Eckstein'),
    '2DOF': ('RIMOSSO', '3D Print'),
    'Gripper Kit': ('RIMOSSO', '3D Print'),
    'Dome Lens': ('DA ORDINARE', 'AliExpress'),
    'Acrylic Dome': ('DA ORDINARE', 'AliExpress'),
}

def find_status(component_name):
    """Find delivery status for a component"""
    name_str = str(component_name)

    # Check Amazon deliveries first
    for keyword, (status, date) in DELIVERY_STATUS.items():
        if keyword.lower() in name_str.lower():
            return status, date

    # Check external orders
    for keyword, (status, note) in EXTERNAL_STATUS.items():
        if keyword.lower() in name_str.lower():
            return status, note

    return None, None

def main():
    print("Loading tracker...")
    wb = openpyxl.load_workbook(TRACKER_PATH)

    # Work with MASTER_TRACKER sheet
    if 'MASTER_TRACKER' not in wb.sheetnames:
        print("ERROR: MASTER_TRACKER sheet not found!")
        return

    ws = wb['MASTER_TRACKER']
    print(f"Working with sheet: MASTER_TRACKER ({ws.max_row} rows)")

    # Add header for delivery columns if not present
    # Find header row
    header_row = None
    for row in range(1, 15):
        cell = ws.cell(row=row, column=2)
        if cell.value and 'Componente' in str(cell.value):
            header_row = row
            # Add delivery status header
            ws.cell(row=row, column=9).value = "Consegna"
            ws.cell(row=row, column=9).font = Font(bold=True)
            ws.cell(row=row, column=10).value = "Data"
            ws.cell(row=row, column=10).font = Font(bold=True)
            break

    updated_count = 0
    items_found = []

    # Valid item markers
    item_markers = ['V', 'v', 'X', 'x']
    emoji_markers = [chr(0x2705), chr(0x2B1C), chr(0x23F3), chr(0x1F4A1), chr(0x2611)]  # Checkbox emojis

    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2)

        if not cell_b.value:
            continue

        # Check if this is an item row (has checkmark/emoji in column A)
        a_val = str(cell_a.value) if cell_a.value else ''

        # Skip header rows
        if 'Componente' in str(cell_b.value):
            continue

        # Check for emoji markers or text markers
        is_item_row = False
        for marker in emoji_markers:
            if marker in a_val:
                is_item_row = True
                break
        if not is_item_row:
            for marker in item_markers:
                if marker in a_val:
                    is_item_row = True
                    break

        if is_item_row:
            component_name = str(cell_b.value)
            status, date = find_status(component_name)

            if status:
                # Update status column (I = 9)
                status_cell = ws.cell(row=row, column=9)
                status_cell.value = status
                status_cell.alignment = Alignment(horizontal='center')
                if status in COLORS:
                    status_cell.fill = COLORS[status]

                # Update date column (J = 10)
                date_cell = ws.cell(row=row, column=10)
                date_cell.value = date
                date_cell.font = Font(italic=True, size=9)

                items_found.append((component_name[:40], status, date))
                updated_count += 1

    # Save
    wb.save(TRACKER_PATH)

    # Print summary
    print(f"\nUpdated {updated_count} items")
    print("\n" + "="*60)
    print("DELIVERY STATUS SUMMARY - 2026-01-14")
    print("="*60)

    ricevuto = [i for i in items_found if i[1] == 'RICEVUTO']
    in_arrivo = [i for i in items_found if i[1] == 'IN ARRIVO']
    da_ordinare = [i for i in items_found if i[1] == 'DA ORDINARE']
    reso = [i for i in items_found if i[1] == 'RESO']
    rimosso = [i for i in items_found if i[1] == 'RIMOSSO']

    print(f"\nRICEVUTO ({len(ricevuto)} items):")
    for name, _, date in ricevuto[:15]:
        print(f"  [OK] {name} - {date}")
    if len(ricevuto) > 15:
        print(f"  ... and {len(ricevuto) - 15} more")

    print(f"\nIN ARRIVO ({len(in_arrivo)} items):")
    for name, _, date in in_arrivo:
        print(f"  [..] {name} - ETA: {date}")

    print(f"\nDA ORDINARE ({len(da_ordinare)} items):")
    for name, _, note in da_ordinare:
        print(f"  [!!] {name} - {note}")

    print(f"\nRESO ({len(reso)} items):")
    for name, _, _ in reso:
        print(f"  [XX] {name}")

    print(f"\nRIMOSSO ({len(rimosso)} items):")
    for name, _, note in rimosso:
        print(f"  [--] {name} - {note}")

    print("\n" + "="*60)
    print(f"File saved: {TRACKER_PATH}")

if __name__ == "__main__":
    main()
