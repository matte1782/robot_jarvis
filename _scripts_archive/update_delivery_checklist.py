"""
OPENDUCK V3 - Delivery Checklist Manager
=========================================
Adds delivery status tracking to OPENDUCK_V3_FINAL_TRACKER.xlsx

Status Legend:
- RICEVUTO (Green): Item received and in hand
- IN ARRIVO (Yellow): Ordered, awaiting delivery
- DA ORDINARE (Red): Not yet ordered
- RIMOSSO (Gray): Removed from order

Usage:
    python update_delivery_checklist.py

Then follow prompts to mark items as received.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import sys

# File paths
TRACKER_PATH = r"C:\Users\matte\Desktop\Desktop OLD\AI\Università AI\courses\personal_project\robot_jarvis\OPENDUCK_V3_FINAL_TRACKER.xlsx"

# Color definitions
COLORS = {
    'RICEVUTO': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Light green
    'IN ARRIVO': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),  # Gold/Yellow
    'DA ORDINARE': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),  # Light red
    'RIMOSSO': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),  # Light gray
    'HEADER': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),  # Blue header
}

def load_tracker():
    """Load the Excel tracker"""
    try:
        wb = openpyxl.load_workbook(TRACKER_PATH)
        return wb
    except FileNotFoundError:
        print(f"ERROR: Tracker not found at {TRACKER_PATH}")
        sys.exit(1)

def find_all_items(ws):
    """Find all orderable items in the tracker"""
    items = []
    for row in range(1, ws.max_row + 1):
        # Check column B (usually component name)
        cell_b = ws.cell(row=row, column=2)
        cell_a = ws.cell(row=row, column=1)

        if cell_b.value and cell_a.value in ['✅', '⬜', '💡', '➕', '⏳']:
            # This is an item row
            item = {
                'row': row,
                'check': str(cell_a.value),
                'name': str(cell_b.value),
                'price': ws.cell(row=row, column=3).value,
                'vendor': ws.cell(row=row, column=4).value,
                'current_status': None
            }

            # Check if there's already a delivery status in column I (9)
            status_cell = ws.cell(row=row, column=9)
            if status_cell.value:
                item['current_status'] = str(status_cell.value).strip()

            items.append(item)

    return items

def add_delivery_column(ws):
    """Add or update the Delivery Status column (Column I)"""
    # Add header if not present
    header_row = None
    for row in range(1, 15):
        cell = ws.cell(row=row, column=2)
        if cell.value and 'Componente' in str(cell.value):
            header_row = row
            break

    if header_row:
        # Set header for delivery column
        delivery_header = ws.cell(row=header_row, column=9)
        delivery_header.value = "📦 Consegna"
        delivery_header.font = Font(bold=True, color="FFFFFF")
        delivery_header.fill = COLORS['HEADER']
        delivery_header.alignment = Alignment(horizontal='center')

def update_item_status(ws, row, status):
    """Update delivery status for a specific row"""
    cell = ws.cell(row=row, column=9)
    cell.value = status
    cell.alignment = Alignment(horizontal='center')

    if status in COLORS:
        cell.fill = COLORS[status]

    # Also update the timestamp in column J
    timestamp_cell = ws.cell(row=row, column=10)
    if status == 'RICEVUTO':
        timestamp_cell.value = datetime.now().strftime('%d/%m/%Y')
        timestamp_cell.font = Font(italic=True, size=9)

def create_delivery_summary_sheet(wb, items):
    """Create or update a delivery summary sheet"""
    # Remove existing summary sheet if present
    if 'DELIVERY_STATUS' in wb.sheetnames:
        del wb['DELIVERY_STATUS']

    # Create new summary sheet
    summary = wb.create_sheet('DELIVERY_STATUS', 0)

    # Header
    summary.cell(row=1, column=1).value = "📦 OPENDUCK V3 - DELIVERY CHECKLIST"
    summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    summary.merge_cells('A1:E1')

    summary.cell(row=2, column=1).value = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    summary.cell(row=2, column=1).font = Font(italic=True)

    # Stats
    received = sum(1 for i in items if i.get('current_status') == 'RICEVUTO')
    in_transit = sum(1 for i in items if i.get('current_status') == 'IN ARRIVO')
    to_order = sum(1 for i in items if i.get('current_status') == 'DA ORDINARE')
    removed = sum(1 for i in items if i.get('current_status') == 'RIMOSSO')

    summary.cell(row=4, column=1).value = "STATUS SUMMARY"
    summary.cell(row=4, column=1).font = Font(bold=True)

    summary.cell(row=5, column=1).value = f"✅ RICEVUTO (In Hand):"
    summary.cell(row=5, column=2).value = received
    summary.cell(row=5, column=2).fill = COLORS['RICEVUTO']

    summary.cell(row=6, column=1).value = f"📬 IN ARRIVO (Shipped):"
    summary.cell(row=6, column=2).value = in_transit
    summary.cell(row=6, column=2).fill = COLORS['IN ARRIVO']

    summary.cell(row=7, column=1).value = f"⏳ DA ORDINARE (To Order):"
    summary.cell(row=7, column=2).value = to_order
    summary.cell(row=7, column=2).fill = COLORS['DA ORDINARE']

    summary.cell(row=8, column=1).value = f"❌ RIMOSSO (Removed):"
    summary.cell(row=8, column=2).value = removed
    summary.cell(row=8, column=2).fill = COLORS['RIMOSSO']

    # Items breakdown by status
    row_num = 10

    # RICEVUTO items
    summary.cell(row=row_num, column=1).value = "✅ ITEMS RECEIVED"
    summary.cell(row=row_num, column=1).font = Font(bold=True, color="006400")
    row_num += 1

    for item in items:
        if item.get('current_status') == 'RICEVUTO':
            summary.cell(row=row_num, column=1).value = f"  • {item['name']}"
            summary.cell(row=row_num, column=2).value = item.get('vendor', '')
            row_num += 1

    if received == 0:
        summary.cell(row=row_num, column=1).value = "  (none yet)"
        summary.cell(row=row_num, column=1).font = Font(italic=True)
        row_num += 1

    row_num += 1

    # IN ARRIVO items
    summary.cell(row=row_num, column=1).value = "📬 ITEMS IN TRANSIT"
    summary.cell(row=row_num, column=1).font = Font(bold=True, color="B8860B")
    row_num += 1

    for item in items:
        if item.get('current_status') == 'IN ARRIVO':
            summary.cell(row=row_num, column=1).value = f"  • {item['name']}"
            summary.cell(row=row_num, column=2).value = item.get('vendor', '')
            row_num += 1

    row_num += 1

    # DA ORDINARE items
    summary.cell(row=row_num, column=1).value = "⏳ ITEMS TO ORDER"
    summary.cell(row=row_num, column=1).font = Font(bold=True, color="8B0000")
    row_num += 1

    for item in items:
        if item.get('current_status') == 'DA ORDINARE':
            summary.cell(row=row_num, column=1).value = f"  • {item['name']}"
            summary.cell(row=row_num, column=2).value = item.get('vendor', '')
            row_num += 1

    # Adjust column widths
    summary.column_dimensions['A'].width = 50
    summary.column_dimensions['B'].width = 20

    return summary

def interactive_update(wb, ws, items):
    """Interactive mode to update delivery status"""
    print("\n" + "="*60)
    print("📦 OPENDUCK V3 - DELIVERY CHECKLIST MANAGER")
    print("="*60)
    print("\nStatus options:")
    print("  1 = RICEVUTO (Received/In Hand)")
    print("  2 = IN ARRIVO (Ordered/In Transit)")
    print("  3 = DA ORDINARE (Not Yet Ordered)")
    print("  4 = RIMOSSO (Removed from Order)")
    print("  s = Skip")
    print("  q = Quit and Save")
    print("\n")

    status_map = {
        '1': 'RICEVUTO',
        '2': 'IN ARRIVO',
        '3': 'DA ORDINARE',
        '4': 'RIMOSSO'
    }

    changes_made = 0

    for item in items:
        current = item.get('current_status', 'Unknown')
        print(f"\n📦 {item['name'][:50]}")
        print(f"   Vendor: {item.get('vendor', 'N/A')}")
        print(f"   Current Status: {current}")

        choice = input("   New status [1/2/3/4/s/q]: ").strip().lower()

        if choice == 'q':
            break
        elif choice == 's' or choice == '':
            continue
        elif choice in status_map:
            new_status = status_map[choice]
            update_item_status(ws, item['row'], new_status)
            item['current_status'] = new_status
            print(f"   → Updated to: {new_status}")
            changes_made += 1

    return changes_made

def batch_update_amazon(wb, ws, items):
    """Batch update all Amazon.it items as IN ARRIVO (ordered)"""
    changes = 0
    for item in items:
        if item.get('vendor') == 'Amazon.it' and item.get('current_status') != 'RICEVUTO':
            update_item_status(ws, item['row'], 'IN ARRIVO')
            item['current_status'] = 'IN ARRIVO'
            changes += 1
    return changes

def main():
    print("Loading tracker...")
    wb = load_tracker()
    ws = wb.active

    print("Finding items...")
    items = find_all_items(ws)
    print(f"Found {len(items)} items in tracker")

    print("Adding delivery column...")
    add_delivery_column(ws)

    # Set default status for items without status
    for item in items:
        if not item.get('current_status'):
            # Default based on current marker
            if '⚠️' in str(ws.cell(row=item['row'], column=7).value or ''):
                update_item_status(ws, item['row'], 'RIMOSSO')
                item['current_status'] = 'RIMOSSO'
            elif item.get('vendor') == 'Amazon.it':
                update_item_status(ws, item['row'], 'IN ARRIVO')
                item['current_status'] = 'IN ARRIVO'
            else:
                update_item_status(ws, item['row'], 'DA ORDINARE')
                item['current_status'] = 'DA ORDINARE'

    # Create summary sheet
    print("Creating delivery summary sheet...")
    create_delivery_summary_sheet(wb, items)

    # Save
    wb.save(TRACKER_PATH)
    print(f"\n✅ Tracker updated: {TRACKER_PATH}")

    # Print summary
    received = sum(1 for i in items if i.get('current_status') == 'RICEVUTO')
    in_transit = sum(1 for i in items if i.get('current_status') == 'IN ARRIVO')
    to_order = sum(1 for i in items if i.get('current_status') == 'DA ORDINARE')
    removed = sum(1 for i in items if i.get('current_status') == 'RIMOSSO')

    print("\n" + "="*50)
    print("📊 DELIVERY STATUS SUMMARY")
    print("="*50)
    print(f"  ✅ RICEVUTO (In Hand):    {received}")
    print(f"  📬 IN ARRIVO (Shipped):   {in_transit}")
    print(f"  ⏳ DA ORDINARE (To Order): {to_order}")
    print(f"  ❌ RIMOSSO (Removed):      {removed}")
    print("="*50)

if __name__ == "__main__":
    main()
