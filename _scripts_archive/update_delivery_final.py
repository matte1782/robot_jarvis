"""
OPENDUCK V3 - Final Delivery Status Update
Based on Amazon order history shared 2026-01-14
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

TRACKER_PATH = r"C:\Users\matte\Desktop\Desktop OLD\AI\Università AI\courses\personal_project\robot_jarvis\OPENDUCK_V3_FINAL_TRACKER.xlsx"

# Status colors
COLORS = {
    'RICEVUTO': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'IN ARRIVO': PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    'DA ORDINARE': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    'RESO': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
}

# Delivery status based on Amazon order history (2026-01-14)
DELIVERY_STATUS = {
    # RICEVUTO (Delivered)
    'QIDI X-Max 3': ('RICEVUTO', 'Dic 2025'),
    'MG90S': ('RICEVUTO', '13/01'),
    'Raspberry Pi 4': ('RICEVUTO', '14/01'),
    'PCA9685': ('RICEVUTO', '15/01'),  # Tomorrow
    'USB-C': ('RICEVUTO', '15/01'),  # Tomorrow
    'Alluminio': ('RICEVUTO', '15/01'),  # Aluminum case
    'Interruttore ON/OFF': ('RICEVUTO', '13/01'),
    'Porta batteria': ('RICEVUTO', '13/01'),
    'batteria 18650': ('RICEVUTO', '13/01'),
    'Limit Switch': ('RICEVUTO', '13/01'),
    'KW11': ('RICEVUTO', '13/01'),
    'WS2812': ('RICEVUTO', '14/01'),
    'NeoPixel': ('RICEVUTO', '14/01'),
    'XT30': ('RICEVUTO', '13-14/01'),
    'YIXISI': ('RICEVUTO', '14/01'),
    'Yiqigou': ('RICEVUTO', '13/01'),
    'HC-SR04': ('RICEVUTO', '13/01'),
    'Ultrasonic': ('RICEVUTO', '13/01'),
    'Silicone': ('RICEVUTO', '13/01'),  # Silicone wire
    'Gruiqrd': ('RICEVUTO', '13/01'),
    'eSUN PLA': ('RICEVUTO', '13/01'),
    'MAX98357': ('RICEVUTO', '13/01'),
    'UBEC': ('RICEVUTO', '13/01'),
    'ZHITING': ('RICEVUTO', '13/01'),
    'ELEGOO': ('RICEVUTO', '14/01'),  # Jumper wires
    'Polymaker': ('RICEVUTO', '14/01'),
    'YINETTECH': ('RICEVUTO', '14/01'),  # Servo horns
    'Servo Braccio': ('RICEVUTO', '14/01'),
    'saldatore': ('RICEVUTO', '14/01'),
    'Viti Cilindriche': ('RICEVUTO', '14/01'),
    'SUNLU': ('RICEVUTO', '14/01'),
    'Silk': ('RICEVUTO', '14/01'),
    'BMS': ('RICEVUTO', '13/01'),
    'TECNOIOT': ('RICEVUTO', '13/01'),
    'Kapton': ('RICEVUTO', '13/01'),
    'YUVKIN': ('RICEVUTO', '13/01'),
    'TPU': ('RICEVUTO', '13/01'),
    'JAYO': ('RICEVUTO', '13/01'),
    'HUAZIZ': ('RICEVUTO', '13/01'),  # Servo extensions
    'Isopropanol': ('RICEVUTO', '14/01'),
    'EQM': ('RICEVUTO', '14/01'),
    'Cuscinetti': ('RICEVUTO', '14/01'),
    'MR63ZZ': ('RICEVUTO', '14/01'),
    'Prusament': ('RICEVUTO', '14/01'),
    'Galaxy': ('RICEVUTO', '14/01'),
    'Enerpower': ('RICEVUTO', '14/01'),  # Battery charger
    'Charger': ('RICEVUTO', '14/01'),

    # IN ARRIVO (In Transit)
    'TXS0108': ('IN ARRIVO', '22/01'),
    'Level Shifter': ('IN ARRIVO', '22/01'),
    'SanDisk': ('IN ARRIVO', '20/01'),
    'microSD': ('IN ARRIVO', '20/01'),
    'Paradisetronic': ('IN ARRIVO', '20/01'),
    'Speaker': ('IN ARRIVO', '20/01'),
    'Altoparlante': ('IN ARRIVO', '20/01'),
    'BNO085': ('IN ARRIVO', '19-22/01'),
    'Adafruit': ('IN ARRIVO', '19-22/01'),
    'IMU': ('IN ARRIVO', '19-22/01'),
    'INMP441': ('IN ARRIVO', '15/01'),
    'AYWHP': ('IN ARRIVO', '15/01'),
    'ruthex': ('IN ARRIVO', '23/01'),
    'Heat Set': ('IN ARRIVO', '23/01'),
    'Flux Pen': ('IN ARRIVO', '6-12/02'),
    'FILO STAGNO': ('IN ARRIVO', '20/01'),
    'ETOPARS': ('IN ARRIVO', '15/01'),
    'Guaina': ('IN ARRIVO', '15/01'),

    # RESO (Returned)
    'Pi Zero': ('RESO', 'Reso'),
    'Zero 2W': ('RESO', 'Reso'),
    'Micro USB': ('RESO', 'Reso'),
    'PowerFast': ('RESO', 'Reso'),
    'Camera Cable': ('RESO', 'Reso'),  # If returned
}

# External orders (not Amazon)
EXTERNAL_STATUS = {
    'Molicel': ('DA ORDINARE', 'Vape Shop'),
    'P30B': ('DA ORDINARE', 'Vape Shop'),
    '18650': ('DA ORDINARE', 'Vape Shop'),
    'AI Camera': ('DA ORDINARE', 'Pimoroni'),
    'IMX500': ('DA ORDINARE', 'Pimoroni'),
    'FE-URT': ('DA ORDINARE', 'AliExpress'),
    'USB-UART': ('DA ORDINARE', 'AliExpress'),
    'STS3215': ('DA ORDINARE', 'Eckstein'),
    'Feetech': ('DA ORDINARE', 'Eckstein'),
    'Gripper': ('RIMOSSO', '3D Print'),
    '2DOF': ('RIMOSSO', '3D Print'),
    'Dome Lens': ('DA ORDINARE', 'AliExpress'),
}

def find_status(component_name):
    """Find delivery status for a component"""
    name_lower = str(component_name).lower()

    # Check Amazon deliveries first
    for keyword, (status, date) in DELIVERY_STATUS.items():
        if keyword.lower() in name_lower:
            return status, date

    # Check external orders
    for keyword, (status, note) in EXTERNAL_STATUS.items():
        if keyword.lower() in name_lower:
            return status, note

    return None, None

def main():
    print("Loading tracker...")
    wb = openpyxl.load_workbook(TRACKER_PATH)
    ws = wb.active

    # Add/update delivery columns
    # Column I = Delivery Status
    # Column J = Delivery Date/Note

    updated_count = 0
    items_found = []

    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2)
        cell_a = ws.cell(row=row, column=1)

        if cell_b.value and cell_a.value in ['OK', '(OK)', '(X)', '?', '!', '*', '>', '>>>', '<<<', '=', '+', '-', '~', '#', '@', '&', '%', '$', '^', '|', '\\', '/', ':', ';', ',', '.', "'", '"', '`', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z']:
            continue

        if cell_b.value and str(cell_a.value) in ['V', 'X']:
            status, date = find_status(cell_b.value)
            if status:
                # Update status column (I)
                status_cell = ws.cell(row=row, column=9)
                status_cell.value = status
                status_cell.alignment = Alignment(horizontal='center')
                if status in COLORS:
                    status_cell.fill = COLORS[status]

                # Update date column (J)
                date_cell = ws.cell(row=row, column=10)
                date_cell.value = date
                date_cell.font = Font(italic=True, size=9)

                items_found.append((cell_b.value[:40], status, date))
                updated_count += 1

    # Also check rows with checkmark symbols
    for row in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row, column=2)
        cell_a = ws.cell(row=row, column=1)

        if not cell_b.value:
            continue

        # Skip if already processed
        if ws.cell(row=row, column=9).value:
            continue

        check_val = str(cell_a.value) if cell_a.value else ''
        if check_val in ['V', 'v', 'x', 'X', '?', '!'] or 'v' in check_val.lower():
            status, date = find_status(cell_b.value)
            if status:
                status_cell = ws.cell(row=row, column=9)
                status_cell.value = status
                status_cell.alignment = Alignment(horizontal='center')
                if status in COLORS:
                    status_cell.fill = COLORS[status]

                date_cell = ws.cell(row=row, column=10)
                date_cell.value = date
                date_cell.font = Font(italic=True, size=9)

                items_found.append((str(cell_b.value)[:40], status, date))
                updated_count += 1

    # Save
    wb.save(TRACKER_PATH)

    # Print summary
    print(f"\nUpdated {updated_count} items")
    print("\n" + "="*60)
    print("DELIVERY STATUS SUMMARY - 2026-01-14")
    print("="*60)

    ricevuto = [i for i in items_found if i[1] == 'RICEVUTO']
    in_arrivo = [i for i in items_found if i[1] == 'IN ARRIVO']
    da_ordinare = [i for i in items_found if i[1] == 'DA ORDINARE']
    reso = [i for i in items_found if i[1] == 'RESO']

    print(f"\nRICEVUTO ({len(ricevuto)} items):")
    for name, _, date in ricevuto[:10]:
        print(f"  [OK] {name} - {date}")
    if len(ricevuto) > 10:
        print(f"  ... and {len(ricevuto) - 10} more")

    print(f"\nIN ARRIVO ({len(in_arrivo)} items):")
    for name, _, date in in_arrivo:
        print(f"  [..] {name} - ETA: {date}")

    print(f"\nDA ORDINARE ({len(da_ordinare)} items):")
    for name, _, note in da_ordinare:
        print(f"  [!!] {name} - {note}")

    print(f"\nRESO ({len(reso)} items):")
    for name, _, _ in reso:
        print(f"  [XX] {name}")

    print("\n" + "="*60)
    print(f"File saved: {TRACKER_PATH}")

if __name__ == "__main__":
    main()
