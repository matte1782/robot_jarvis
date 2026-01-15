from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

wb = Workbook()
ws = wb.active
ws.title = 'MASTER_TRACKER'

# Title
ws['A1'] = 'OPENDUCK MINI V3 ENHANCED - MASTER PROCUREMENT TRACKER'
ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells('A1:H1')
ws.row_dimensions[1].height = 35

ws['A2'] = f'Last Update: {datetime.now().strftime("%d %B %Y, %H:%M")} | Budget Totale: ~€1,718'
ws['A2'].font = Font(size=11, italic=True, color='FFFFFF')
ws['A2'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws['A2'].alignment = Alignment(horizontal='center')
ws.merge_cells('A2:H2')
ws.row_dimensions[2].height = 20

# Column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 25

row = 4

# SECTION 1 - ALREADY PURCHASED
ws[f'A{row}'] = '✅ FASE 1 - GIA ACQUISTATO'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')
ws.row_dimensions[row].height = 25
row += 1

headers = ['✓', 'Componente', 'Prezzo €', 'Fornitore', 'Data Ordine', 'Note', 'Status', 'Tracking']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
row += 1

# Already purchased item
already_purchased = [
    ('✅', 'QIDI X-Max 3 Stampante 3D', 599.00, 'Amazon.it', 'Dic 2025', 'Core System - 300x300x300mm build', 'RICEVUTO', 'Consegnato'),
]

for item in already_purchased:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row, col, value)
        if col == 3:
            cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    row += 1

ws[f'C{row}'] = 'TOTALE GIA SPESO:'
ws[f'C{row}'].font = Font(bold=True)
ws[f'D{row}'] = 599.00
ws[f'D{row}'].font = Font(bold=True)
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
row += 2

# SECTION 2 - AMAZON.IT ORDER (32 items)
ws[f'A{row}'] = '🛒 FASE 2 - ORDINE AMAZON.IT (PROCEDI CHECKOUT)'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')
ws.row_dimensions[row].height = 25
row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
row += 1

amazon_start = row

# Amazon.it items (32 items from cart)
amazon_items = [
    ('⬜', 'FILAMENTI - Prusament Galaxy PLA 2×1kg', 75.98, 'Amazon.it', 'DA ORDINARE', 'Premium structural filament', 'CARRELLO', '13-16 gen'),
    ('⬜', 'FILAMENTI - Polymaker PLA Pro Bianco 1kg', 37.99, 'Amazon.it', 'DA ORDINARE', 'Engineering-grade PLA', 'CARRELLO', '13-16 gen'),
    ('⬜', 'FILAMENTI - eSUN PLA+ Bianco 1kg', 20.52, 'Amazon.it', 'DA ORDINARE', 'Reliable standard PLA', 'CARRELLO', '14 gen'),
    ('⬜', 'FILAMENTI - SUNLU Silk PLA Plus Triplo 1kg', 18.99, 'Amazon.it', 'DA ORDINARE', 'Creative Nero-Oro-Viola', 'CARRELLO', '13-16 gen'),
    ('⬜', 'FILAMENTI - JAYO TPU 95A 0.5kg', 15.99, 'Amazon.it', 'DA ORDINARE', 'Flexible feet pads', 'CARRELLO', '13 gen'),
    ('⬜', 'Raspberry Pi Zero 2W 512MB RAM', 23.80, 'Amazon.it', 'DA ORDINARE', 'Primary compute unit', 'CARRELLO', '13-16 gen'),
    ('⬜', 'SanDisk Ultra 32GB microSD Class 10', 11.99, 'Amazon.it', 'DA ORDINARE', 'OS storage A1 rated', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Adafruit BNO085 9-DOF IMU', 42.59, 'Amazon.it', 'DA ORDINARE', 'Motion tracking sensor', 'CARRELLO', '13-22 gen'),
    ('⬜', 'AYWHP INMP441 I2S Microphone (6 pcs)', 15.99, 'Amazon.it', 'DA ORDINARE', 'Audio input system', 'CARRELLO', '13-16 gen'),
    ('⬜', 'AZDelivery MAX98357A I2S Amplifier', 10.99, 'Amazon.it', 'DA ORDINARE', 'Audio output amplifier', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Paradisetronic Speaker 2W 8Ω 40mm', 10.05, 'Amazon.it', 'DA ORDINARE', 'Mini speaker for voice', 'CARRELLO', '16 gen'),
    ('⬜', 'Aihasd HC-SR04 Ultrasonic (3 pcs)', 7.59, 'Amazon.it', 'DA ORDINARE', 'Obstacle detection', 'CARRELLO', '13-16 gen'),
    ('⬜', 'TXS0108E Level Shifter (10 pcs)', 9.99, 'Amazon.it', 'DA ORDINARE', '3.3V ↔ 5V bidirectional', 'CARRELLO', '22 gen'),
    ('⬜', '5pcs 2S BMS 20A Battery Protection', 17.18, 'Amazon.it', 'DA ORDINARE', 'Li-ion battery safety', 'CARRELLO', '13-16 gen'),
    ('⬜', 'ZHITING UBEC 5V 3A Step-Down', 6.99, 'Amazon.it', 'DA ORDINARE', 'Electronics power rail', 'CARRELLO', '13-16 gen'),
    ('⬜', 'YIXISI 5 Paia XT30 Connectors', 9.59, 'Amazon.it', 'DA ORDINARE', 'Power distribution 30A', 'CARRELLO', '13 gen'),
    ('⬜', 'Yiqigou 2pcs XT30 Extension Cable', 8.79, 'Amazon.it', 'DA ORDINARE', 'Power cable extensions', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Enerpower 2S Li-ion Charger 7.4V', 19.44, 'Amazon.it', 'DA ORDINARE', 'Battery charging system', 'CARRELLO', '13 gen'),
    ('⬜', 'ruthex M3 Heat Set Inserts 100pcs', 10.99, 'Amazon.it', 'DA ORDINARE', 'M3x5x4mm brass inserts', 'CARRELLO', '17 gen'),
    ('⬜', 'Viti Cilindriche M2 M3 M4 1080pcs', 10.79, 'Amazon.it', 'DA ORDINARE', 'Socket head cap screws', 'CARRELLO', '13-16 gen'),
    ('⬜', '10 mini Cuscinetti MR63ZZ 3x6x2.5mm', 8.35, 'Amazon.it', 'DA ORDINARE', 'Ball bearings for joints', 'CARRELLO', '16 gen'),
    ('⬜', 'ELEGOO 120pcs Jumper Wire Kit', 9.49, 'Amazon.it', 'DA ORDINARE', 'Sensor wiring Dupont', 'CARRELLO', '13-16 gen'),
    ('⬜', 'AZDelivery Pi Zero Camera Cable 15cm', 6.99, 'Amazon.it', 'DA ORDINARE', 'Flex ribbon for camera', 'CARRELLO', '13-16 gen'),
    ('⬜', 'AZDelivery 5x MG90S Metal Gear Servo', 23.99, 'Amazon.it', 'DA ORDINARE', 'Arm servos 13g (4+1)', 'CARRELLO', '13-16 gen'),
    ('⬜', 'YINETTECH 4 Pezzi 25T Servo Horn', 12.09, 'Amazon.it', 'DA ORDINARE', 'Aluminum servo horns', 'CARRELLO', '13 gen'),
    ('⬜', 'HUAZIZ 24 Pezzi Servo Extension', 21.98, 'Amazon.it', 'DA ORDINARE', 'Servo cables (2 packs)', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Set Saldatore 60W Temperature Control', 18.99, 'Amazon.it', 'DA ORDINARE', 'Soldering station kit', 'CARRELLO', '13-16 gen'),
    ('⬜', 'FILO STAGNO 60/40 100gr 0.8mm', 7.50, 'Amazon.it', 'DA ORDINARE', 'Leaded solder ✅ CORRECT', 'CARRELLO', '16 gen'),
    ('⬜', 'Flux Pen 951 10ml No-Clean', 3.87, 'Amazon.it', 'DA ORDINARE', 'Soldering flux pen', 'CARRELLO', '6-12 feb'),
    ('⬜', 'EQM Isopropanolo 99.9% 1L', 10.20, 'Amazon.it', 'DA ORDINARE', 'Electronics cleaning ✅', 'CARRELLO', '13 gen'),
    ('⬜', 'YUVKIN 5 Pezzi Nastro Kapton', 7.99, 'Amazon.it', 'DA ORDINARE', 'Heat-resistant tape set', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Gruiqrd Filo Silicone 16 gauge 5m', 7.99, 'Amazon.it', 'DA ORDINARE', 'Power wiring silicone', 'CARRELLO', '13-16 gen'),
]

for item in amazon_items:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row, col, value)
        if col == 3:
            cell.number_format = '#,##0.00'
        if 'CORRECT' in str(item[5]) or '✅' in str(item[5]):
            cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    row += 1

ws[f'C{row}'] = 'TOTALE AMAZON #1:'
ws[f'C{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'] = f'=SUM(D{amazon_start}:D{row-1})'
ws[f'D{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ws[f'G{row}'] = 'ORDINATO 12/01'
ws[f'G{row}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ws[f'G{row}'].font = Font(bold=True, color='006100')
amazon1_total_row = row
row += 2

# SECTION 2B - AMAZON.IT ORDER #2 (5 items - Supplementary)
ws[f'A{row}'] = '🛒 FASE 2B - AMAZON.IT ORDINE #2 (DA ORDINARE ORA)'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')
ws.row_dimensions[row].height = 25
row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
row += 1

amazon2_start = row

amazon2_items = [
    ('⬜', 'CABLEPELADO Porta batteria 18650 2S 7.4V', 5.20, 'Amazon.it', 'DA ORDINARE', 'Battery holder for 2x 18650', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Mini Interruttore ON/OFF 5pcs 10A/125V', 5.68, 'Amazon.it', 'DA ORDINARE', 'Power switch for robot', 'CARRELLO', '13-16 gen'),
    ('⬜', 'Taiss Micro Limit Switch 10pcs KW11-3Z-02', 10.44, 'Amazon.it', 'DA ORDINARE', 'Foot contact sensors (need 4)', 'CARRELLO', '13-16 gen'),
    ('⬜', 'ETOPARS 3mm Guaina Cavi Intrecciata 10m', 8.79, 'Amazon.it', 'DA ORDINARE', 'Cable management sleeve', 'CARRELLO', '16 gen'),
    ('⬜', 'Amazon PowerFast Micro USB 1.5m', 9.99, 'Amazon.it', 'DA ORDINARE', 'USB-A to MicroUSB for Pi Zero', 'CARRELLO', '13-16 gen'),
]

for item in amazon2_items:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row, col, value)
        if col == 3:
            cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    row += 1

ws[f'C{row}'] = 'TOTALE AMAZON #2:'
ws[f'C{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'] = f'=SUM(D{amazon2_start}:D{row-1})'
ws[f'D{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
amazon2_total_row = row
row += 2

# SECTION 3 - EXTERNAL ORDERS PRIORITY
ws[f'A{row}'] = '🌍 FASE 3 - ORDINI ESTERNI PRIORITARI (ORDINA OGGI)'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')
ws.row_dimensions[row].height = 25
row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
row += 1

external_start = row

external_items = [
    ('⬜', '2× Molicel INR18650-P30B 3000mAh', 15.00, 'Vape Shop Monza', 'OGGI', 'Ritiro immediato - controlla QR!', 'DA ORDINARE', 'Stesso giorno'),
    ('⬜', 'Raspberry Pi AI Camera IMX500 12MP', 78.00, 'Pimoroni.uk', 'OGGI', 'On-sensor AI object detection', 'DA ORDINARE', '5-7 giorni UK'),
    ('⬜', 'FE-URT-1 USB-UART Servo Controller', 12.90, 'AliExpress', 'OGGI', 'Half-duplex serial bus', 'DA ORDINARE', '15-25 giorni'),
]

for item in external_items:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row, col, value)
        if col == 3:
            cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    row += 1

ws[f'C{row}'] = 'TOTALE ESTERNI:'
ws[f'C{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'] = f'=SUM(D{external_start}:D{row-1})'
ws[f'D{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
external_total_row = row
row += 2

# SECTION 4 - WAITING FOR QUOTE
ws[f'A{row}'] = '⏳ FASE 4 - ATTESA CONFERMA FORNITORE (NON ORDINARE)'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')
ws.row_dimensions[row].height = 25
row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
row += 1

waiting_start = row

waiting_items = [
    ('⏳', 'Email quotazione a info@eckstein-shop.de', 0.00, 'Eckstein.de', 'OGGI', 'Richiesta prezzo 16× STS3215', 'DA INVIARE', 'Risposta 24-48h'),
    ('⏳', '16× Feetech STS3215 Servo 7.4V 19kg·cm', 400.00, 'Eckstein.de', 'Dopo quote', 'ATTENDI conferma disponibilità', 'ATTESA RISPOSTA', 'Stima 7 giorni'),
]

for item in waiting_items:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row, col, value)
        if col == 3:
            cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        cell.font = Font(color='BF8F00')
    row += 1

ws[f'C{row}'] = 'STIMA ECKSTEIN:'
ws[f'C{row}'].font = Font(bold=True, size=11, color='BF8F00')
ws[f'D{row}'] = 400.00
ws[f'D{row}'].font = Font(bold=True, size=11, color='BF8F00')
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
waiting_total_row = row
row += 2

# SECTION 5 - OPTIONAL
ws[f'A{row}'] = '💡 FASE 5 - COMPONENTI OPZIONALI (NON URGENTE)'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')
ws.row_dimensions[row].height = 25
row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row, col, header)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
row += 1

optional_start = row

optional_items = [
    ('💡', '2DOF Aluminum Robot Gripper Kit', 27.00, 'AliExpress', 'Dopo test', '15cm reach, 50g payload', 'OPZIONALE', '15-30 giorni'),
    ('💡', 'Acrylic Dome Lenses 25mm (2pcs)', 9.00, 'AliExpress', 'Dopo test', 'Cosmetic eye effect', 'OPZIONALE', '15-30 giorni'),
]

for item in optional_items:
    for col, value in enumerate(item, 1):
        cell = ws.cell(row, col, value)
        if col == 3:
            cell.number_format = '#,##0.00'
        cell.fill = PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid')
    row += 1

ws[f'C{row}'] = 'TOTALE OPZIONALI:'
ws[f'C{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'] = f'=SUM(D{optional_start}:D{row-1})'
ws[f'D{row}'].font = Font(bold=True, size=11)
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
optional_total_row = row
row += 2

# GRAND TOTAL
ws[f'A{row}'] = 'BUDGET TOTALE PROGETTO OPENDUCK V3'
ws[f'A{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'A{row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws.merge_cells(f'A{row}:C{row}')
ws[f'D{row}'] = f'=D{amazon1_total_row}+D{amazon2_total_row}+D{external_total_row}+D{waiting_total_row}+D{optional_total_row}+599'
ws[f'D{row}'].font = Font(bold=True, size=14, color='FFFFFF')
ws[f'D{row}'].number_format = '#,##0.00'
ws[f'D{row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws[f'E{row}'] = '(Include QIDI)'
ws[f'E{row}'].font = Font(italic=True, size=10, color='FFFFFF')
ws[f'E{row}'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws.merge_cells(f'E{row}:H{row}')

# Get script directory for output path
script_dir = os.path.dirname(os.path.abspath(__file__))
output_file = os.path.join(script_dir, 'OPENDUCK_V3_FINAL_TRACKER.xlsx')

# Save
wb.save(output_file)
print('=' * 60)
print('OPENDUCK V3 MASTER TRACKER - CREATO CON SUCCESSO!')
print('=' * 60)
print(f'File: {output_file}')
print('')
print('SEZIONI CREATE:')
print('  1. FASE 1 - Gia Acquistato (QIDI X-Max 3 - €599)')
print('  2. FASE 2A - Amazon.it Ordine #1 (35 items - €483.98) ORDINATO')
print('  2B. FASE 2B - Amazon.it Ordine #2 (5 items - €40.10) DA ORDINARE')
print('  3. FASE 3 - Ordini Esterni (batterie, camera, FE-URT-1)')
print('  4. FASE 4 - Attesa Conferma (16x STS3215 Eckstein)')
print('  5. FASE 5 - Componenti Opzionali')
print('')
print('BUDGET TOTALE: ~€1,718 (formula auto-calcolata)')
print('')
print('LEGENDA STATUS:')
print('  VERDE   = Ordinato/Ricevuto')
print('  GIALLO  = Da ordinare oggi')
print('  ARANCIO = In attesa risposta')
print('  VIOLA   = Opzionale')
print('=' * 60)
