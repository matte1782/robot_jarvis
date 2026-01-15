import sys
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

# Fix Windows console encoding for Unicode characters
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('OPENDUCK_V3_FINAL_TRACKER.xlsx')
ws = wb.active

# Find the Pi Zero 2W row and update it
for row in ws.iter_rows(min_row=2):
    # Check if this is the Pi Zero 2W row
    if row[1].value and 'Raspberry Pi Zero 2W' in str(row[1].value):
        # Update status to show it needs to be removed/returned
        status_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value and 'STATUS' in str(cell.value).upper():
                status_col = idx
                break

        if status_col is not None:
            row[status_col].value = "⚠️ DA RIMUOVERE - CPU insufficiente"
            row[status_col].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # Add note about bottleneck
        note_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value and 'NOTE' in str(cell.value).upper():
                note_col = idx
                break

        if note_col is not None:
            row[note_col].value = "BOTTLENECK: CPU 47-83% carico, rischio RL policy. Upgrade a Pi 4 4GB necessario."

    # Check if this is the micro-USB cable row
    if row[1].value and 'micro' in str(row[1].value).lower() and 'usb' in str(row[1].value).lower() and 'cavo' in str(row[1].value).lower():
        if status_col is not None:
            row[status_col].value = "⚠️ DA RIMUOVERE - Pi 4 usa USB-C"
            row[status_col].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# Find the last row to add new items
last_row = ws.max_row

# Add new items for Pi 4 upgrade
new_items = [
    {
        'name': 'Raspberry Pi 4 Model B 4GB',
        'qty': 1,
        'price': 76.60,
        'source': 'Amazon.it',
        'status': '✅ DA ORDINARE - UPGRADE',
        'note': 'Sostituisce Pi Zero 2W. CPU 3-4x più veloce, elimina rischio bottleneck'
    },
    {
        'name': 'Alimentatore USB-C 5V 3A per Pi 4',
        'qty': 1,
        'price': 13.25,
        'source': 'Amazon.it',
        'status': '✅ DA ORDINARE - NECESSARIO',
        'note': 'Pi 4 richiede USB-C, non micro-USB'
    },
    {
        'name': 'Case Alluminio + Dissipatore Passivo Pi 4',
        'qty': 1,
        'price': 13.19,
        'source': 'Amazon.it',
        'status': '✅ DA ORDINARE - RACCOMANDATO',
        'note': 'Previene thermal throttling a 80°C. Mantiene CPU a 65°C'
    }
]

# Find column indices
col_map = {}
for idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        header = str(cell.value).upper()
        if 'COMPONENTE' in header or 'NOME' in header or 'ITEM' in header:
            col_map['name'] = idx
        elif 'QTY' in header or 'QTÀ' in header or 'QUANTITÀ' in header:
            col_map['qty'] = idx
        elif 'PREZZO' in header or 'PRICE' in header or 'COSTO' in header:
            col_map['price'] = idx
        elif 'SOURCE' in header or 'FORNITORE' in header:
            col_map['source'] = idx
        elif 'STATUS' in header or 'STATO' in header:
            col_map['status'] = idx
        elif 'NOTE' in header or 'NOTES' in header:
            col_map['note'] = idx

# Add new rows
for i, item in enumerate(new_items, start=1):
    new_row = last_row + i
    if 'name' in col_map:
        ws.cell(row=new_row, column=col_map['name']).value = item['name']
    if 'qty' in col_map:
        ws.cell(row=new_row, column=col_map['qty']).value = item['qty']
    if 'price' in col_map:
        ws.cell(row=new_row, column=col_map['price']).value = item['price']
    if 'source' in col_map:
        ws.cell(row=new_row, column=col_map['source']).value = item['source']
    if 'status' in col_map:
        cell = ws.cell(row=new_row, column=col_map['status'])
        cell.value = item['status']
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    if 'note' in col_map:
        ws.cell(row=new_row, column=col_map['note']).value = item['note']

# Add summary section at the bottom
summary_row = last_row + len(new_items) + 2
ws.cell(row=summary_row, column=1).value = "ENGINEERING AUDIT SUMMARY - 2026-01-13"
ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)

summary_row += 1
ws.cell(row=summary_row, column=1).value = "BOTTLENECK TROVATO:"
ws.cell(row=summary_row, column=2).value = "Pi Zero 2W CPU carico 47-83% (single core), potenza 5V rail marginal (1.3A vs 3A limite)"

summary_row += 1
ws.cell(row=summary_row, column=1).value = "SOLUZIONE:"
ws.cell(row=summary_row, column=2).value = "Upgrade a Pi 4 4GB (3-4x più veloce, 4GB RAM vs 512MB)"

summary_row += 1
ws.cell(row=summary_row, column=1).value = "COSTO UPGRADE:"
ws.cell(row=summary_row, column=2).value = "+€69.25 netto (€102.99 nuovi item - €33.79 rimossi)"

summary_row += 1
ws.cell(row=summary_row, column=1).value = "TIMELINE IMPATTO:"
ws.cell(row=summary_row, column=2).value = "ZERO - stampa 3D è bottleneck (40-60h), Pi serve solo Phase 3"

summary_row += 1
ws.cell(row=summary_row, column=1).value = "AZIONE RICHIESTA:"
ws.cell(row=summary_row, column=2).value = "1) Rimuovi Pi Zero 2W + micro-USB da carrello  2) Aggiungi Pi 4 + USB-C + case  3) Ordina"

# Save the updated workbook
wb.save('OPENDUCK_V3_FINAL_TRACKER.xlsx')

print("✅ Tracker aggiornato con successo!")
print(f"   - Pi Zero 2W marcato per rimozione")
print(f"   - Micro-USB marcato per rimozione")
print(f"   - 3 nuovi item aggiunti (Pi 4, USB-C, case)")
print(f"   - Engineering audit summary aggiunto")
