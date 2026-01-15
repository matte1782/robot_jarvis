"""
Update OPENDUCK_V3_FINAL_TRACKER.xlsx
- Remove commercial 2DOF Aluminum Robot Gripper Kit from FASE 5
- Add note about 3D printed gripper approach (APPROVED by technical review)
- Update MG90S servo status to RICEVUTO (received)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# File path
tracker_path = r"C:\Users\matte\Desktop\Desktop OLD\AI\Università AI\courses\personal_project\robot_jarvis\OPENDUCK_V3_FINAL_TRACKER.xlsx"

# Load workbook
wb = openpyxl.load_workbook(tracker_path)

# Get all sheet names
print(f"Sheets found: {wb.sheetnames}")

# Find the main sheet (usually first one or named appropriately)
ws = wb.active
print(f"Active sheet: {ws.title}")

# Scan for the gripper kit row and remove/mark it
gripper_found = False
mg90s_found = False
changes_made = []

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            cell_str = str(cell.value).lower()

            # Find and mark 2DOF Aluminum Robot Gripper Kit as REMOVED
            if "2dof" in cell_str and "gripper" in cell_str:
                gripper_found = True
                # Mark the row as removed
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=c).font = Font(strikethrough=True, color="999999")
                # Add note in status column (usually last used column)
                status_col = ws.max_column
                ws.cell(row=row, column=status_col).value = "RIMOSSO - 3D PRINT"
                changes_made.append(f"Row {row}: Marked 2DOF gripper kit as REMOVED (3D print instead)")
                print(f"Found gripper kit at row {row} - marking as REMOVED")

            # Find MG90S servo row and update status to RICEVUTO
            if "mg90s" in cell_str and "servo" in cell_str:
                mg90s_found = True
                # Find status column and update
                for c in range(1, ws.max_column + 1):
                    status_cell = ws.cell(row=row, column=c)
                    if status_cell.value and "status" in str(status_cell.value).lower():
                        # This is a header row, skip
                        continue
                    if status_cell.value in ["CARRELLO", "DA ORDINARE", "IN ARRIVO", "ORDINATO"]:
                        status_cell.value = "RICEVUTO"
                        status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        changes_made.append(f"Row {row}: Updated MG90S servo status to RICEVUTO")
                        print(f"Updated MG90S status at row {row}")

# Add a new section for 3D PRINTED GRIPPER if space allows
# Find last row and add note
last_row = ws.max_row + 3

ws.cell(row=last_row, column=1).value = "=== 3D PRINTED GRIPPER (APPROVED) ==="
ws.cell(row=last_row, column=1).font = Font(bold=True, color="006400")
ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=5)

last_row += 1
ws.cell(row=last_row, column=1).value = "Component"
ws.cell(row=last_row, column=2).value = "Source"
ws.cell(row=last_row, column=3).value = "Cost"
ws.cell(row=last_row, column=4).value = "Status"
ws.cell(row=last_row, column=5).value = "Notes"
for c in range(1, 6):
    ws.cell(row=last_row, column=c).font = Font(bold=True)

last_row += 1
ws.cell(row=last_row, column=1).value = "MG90S Servos (5x)"
ws.cell(row=last_row, column=2).value = "Amazon.it"
ws.cell(row=last_row, column=3).value = "€23.99"
ws.cell(row=last_row, column=4).value = "RICEVUTO"
ws.cell(row=last_row, column=4).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
ws.cell(row=last_row, column=5).value = "Already purchased & received"

last_row += 1
ws.cell(row=last_row, column=1).value = "3D Printed Gripper (2x)"
ws.cell(row=last_row, column=2).value = "Thingiverse #5149951"
ws.cell(row=last_row, column=3).value = "~€2-5"
ws.cell(row=last_row, column=4).value = "DA STAMPARE"
ws.cell(row=last_row, column=4).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
ws.cell(row=last_row, column=5).value = "Print with PLA+ 50% infill"

last_row += 1
ws.cell(row=last_row, column=1).value = "M3x15mm Bolts + Nyloc Nuts"
ws.cell(row=last_row, column=2).value = "Amazon.it"
ws.cell(row=last_row, column=3).value = "~€3"
ws.cell(row=last_row, column=4).value = "IN ORDINE"
ws.cell(row=last_row, column=4).fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
ws.cell(row=last_row, column=5).value = "May already have in hardware kit"

last_row += 2
ws.cell(row=last_row, column=1).value = f"Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
ws.cell(row=last_row, column=1).font = Font(italic=True, color="666666")

ws.cell(row=last_row+1, column=1).value = "Technical Review: APPROVED for 3D printing (lighter, cheaper, customizable)"
ws.cell(row=last_row+1, column=1).font = Font(italic=True, color="006400")

changes_made.append("Added 3D PRINTED GRIPPER section")

# Save workbook
wb.save(tracker_path)
print(f"\n{'='*50}")
print("TRACKER UPDATED SUCCESSFULLY")
print(f"{'='*50}")
print(f"\nChanges made:")
for change in changes_made:
    print(f"  - {change}")

print(f"\nFile saved: {tracker_path}")
