import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Fix Windows console encoding
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('OPENDUCK_V3_FINAL_TRACKER.xlsx')
ws = wb.active

# Find the FASE 2 section and mark items to remove
fase2_start = None
for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50), 1):
    if row[0].value and 'FASE 2' in str(row[0].value):
        fase2_start = idx + 1  # Next row has headers
        break

if fase2_start:
    # Find items to mark as "DA RIMUOVERE"
    for row_idx in range(fase2_start + 1, fase2_start + 100):
        componente_cell = ws.cell(row=row_idx, column=2)  # Column B = Componente
        status_cell = ws.cell(row=row_idx, column=7)      # Column G = Status
        note_cell = ws.cell(row=row_idx, column=6)        # Column F = Note

        if not componente_cell.value:
            continue

        item_name = str(componente_cell.value).lower()

        # Mark Pi Zero 2W
        if 'raspberry pi zero 2w' in item_name:
            status_cell.value = "⚠️ DA RIMUOVERE"
            status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            status_cell.font = Font(bold=True)
            note_cell.value = "UPGRADE: CPU insufficiente (47-83% carico). Sostituire con Pi 4 4GB"

        # Mark Micro-USB cable
        if 'micro' in item_name and 'usb' in item_name:
            status_cell.value = "⚠️ DA RIMUOVERE"
            status_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            status_cell.font = Font(bold=True)
            note_cell.value = "Pi 4 usa USB-C, non micro-USB"

# Find last row with data in FASE 2
last_data_row = fase2_start + 1
for row_idx in range(fase2_start + 1, fase2_start + 100):
    if ws.cell(row=row_idx, column=2).value:
        last_data_row = row_idx

# Add new items section
new_section_start = last_data_row + 3

# Section header
ws.merge_cells(f'A{new_section_start}:H{new_section_start}')
header_cell = ws.cell(row=new_section_start, column=1)
header_cell.value = "➕ ITEMS DA AGGIUNGERE AL CARRELLO"
header_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
header_cell.font = Font(bold=True, size=12)
header_cell.alignment = Alignment(horizontal='center')

# Column headers for new items
new_section_start += 1
headers = ['✓', 'Componente', 'Prezzo €', 'Fornitore', 'Data Ordine', 'Note', 'Status', 'Tracking']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=new_section_start, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# New items to add
new_items = [
    {
        'check': '➕',
        'componente': 'Raspberry Pi 4 Model B 4GB RAM',
        'prezzo': 76.60,
        'fornitore': 'Amazon.it',
        'data': 'DA ORDINARE',
        'note': '🔄 UPGRADE: 3-4× CPU più veloce, 8× RAM. Elimina bottleneck rilevato',
        'status': 'CARRELLO',
        'tracking': '13-16 gen'
    },
    {
        'check': '➕',
        'componente': 'Alimentatore USB-C 5V 3A per Raspberry Pi 4',
        'prezzo': 13.25,
        'fornitore': 'Amazon.it',
        'data': 'DA ORDINARE',
        'note': '🔄 NECESSARIO: Pi 4 usa USB-C, non micro-USB',
        'status': 'CARRELLO',
        'tracking': '13-16 gen'
    },
    {
        'check': '➕',
        'componente': 'Case Alluminio + Dissipatore Passivo Raspberry Pi 4',
        'prezzo': 13.19,
        'fornitore': 'Amazon.it',
        'data': 'DA ORDINARE',
        'note': '🔄 RACCOMANDATO: Previene thermal throttling (80°C limit). Mantiene 65°C',
        'status': 'CARRELLO',
        'tracking': '13-16 gen'
    },
    {
        'check': '➕',
        'componente': 'PCA9685 PWM Driver Board 16-Ch I2C (2pcs) - GERUI',
        'prezzo': 10.09,
        'fornitore': 'Amazon.it',
        'data': 'DA ORDINARE',
        'note': '⚠️ MANCANTE: Controller I2C per 4× MG90S servos (braccia). Address 0x40',
        'status': 'CARRELLO',
        'tracking': '13-16 gen'
    }
]

# Add new items
for idx, item in enumerate(new_items, 1):
    row_num = new_section_start + idx
    ws.cell(row=row_num, column=1).value = item['check']
    ws.cell(row=row_num, column=2).value = item['componente']
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    ws.cell(row=row_num, column=3).value = item['prezzo']
    ws.cell(row=row_num, column=4).value = item['fornitore']
    ws.cell(row=row_num, column=5).value = item['data']
    ws.cell(row=row_num, column=6).value = item['note']
    ws.cell(row=row_num, column=7).value = item['status']
    ws.cell(row=row_num, column=8).value = item['tracking']

# RECAP section
recap_start = new_section_start + len(new_items) + 3

# Main header
ws.merge_cells(f'A{recap_start}:H{recap_start}')
header_cell = ws.cell(row=recap_start, column=1)
header_cell.value = "📋 RECAP COMPLETO MODIFICHE - ACQUISTO UNICO DOMANI"
header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_cell.font = Font(bold=True, size=14, color="FFFFFF")
header_cell.alignment = Alignment(horizontal='center', vertical='center')

# Step 1: RIMUOVI
recap_start += 2
ws.merge_cells(f'A{recap_start}:H{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "❌ STEP 1: RIMUOVI DAL CARRELLO AMAZON"
section_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
section_cell.font = Font(bold=True, size=11)
section_cell.alignment = Alignment(horizontal='left')

recap_start += 1
remove_items = [
    ["⚠️ Raspberry Pi Zero 2W", "€23.80", "CPU insufficiente (47-83% carico single core)"],
    ["⚠️ Amazon PowerFast Micro USB 1.5m", "€9.99", "Pi 4 non usa micro-USB, usa USB-C"],
    ["", "", ""],
    ["TOTALE DA RIMUOVERE:", "-€33.79", ""]
]

for item in remove_items:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=3).value = item[2]
    if "TOTALE" in item[0]:
        ws.cell(row=recap_start, column=1).font = Font(bold=True)
        ws.cell(row=recap_start, column=2).font = Font(bold=True, color="008000")
    else:
        ws.cell(row=recap_start, column=1).font = Font(color="C00000")
    recap_start += 1

# Step 2: AGGIUNGI
recap_start += 1
ws.merge_cells(f'A{recap_start}:H{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "✅ STEP 2: AGGIUNGI AL CARRELLO AMAZON"
section_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
section_cell.font = Font(bold=True, size=11)
section_cell.alignment = Alignment(horizontal='left')

recap_start += 1
add_items_recap = [
    ["✅ Raspberry Pi 4 Model B 4GB", "€76.60", "3-4× più veloce, 8× RAM vs Pi Zero 2W"],
    ["✅ Alimentatore USB-C 5V 3A", "€13.25", "Alimentazione corretta per Pi 4"],
    ["✅ Case Alluminio + Dissipatore", "€13.19", "Previene throttling termico (80°C)"],
    ["✅ PCA9685 PWM Driver (2pcs)", "€10.09", "Controller I2C per servos braccia"],
    ["", "", ""],
    ["TOTALE DA AGGIUNGERE:", "+€113.13", ""]
]

for item in add_items_recap:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=3).value = item[2]
    if "TOTALE" in item[0]:
        ws.cell(row=recap_start, column=1).font = Font(bold=True)
        ws.cell(row=recap_start, column=2).font = Font(bold=True, color="C00000")
    else:
        ws.cell(row=recap_start, column=1).font = Font(color="008000")
    recap_start += 1

# Cost summary
recap_start += 1
ws.merge_cells(f'A{recap_start}:H{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "💰 RIEPILOGO COSTI FINALI"
section_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
section_cell.font = Font(bold=True, size=11)

recap_start += 1
cost_items = [
    ["Items rimossi:", "-€33.79"],
    ["Items aggiunti:", "+€113.13"],
    ["", ""],
    ["COSTO NETTO UPGRADE:", "+€79.34"]
]

for item in cost_items:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    if "NETTO" in item[0]:
        ws.cell(row=recap_start, column=1).font = Font(bold=True, size=12)
        ws.cell(row=recap_start, column=2).font = Font(bold=True, size=12, color="C00000")
    recap_start += 1

# Why upgrade
recap_start += 1
ws.merge_cells(f'A{recap_start}:H{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "🔍 PERCHÉ QUESTO UPGRADE?"
section_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_cell.font = Font(bold=True, size=11)

recap_start += 1
why_items = [
    ["1. BOTTLENECK CPU:", "Pi Zero 2W carico 47-83% single core (rischio RL policy)"],
    ["2. POTENZA MARGINALE:", "5V rail 1.3A picco vs 3A UBEC limit (con braccia)"],
    ["3. COMPONENTE MANCANTE:", "PCA9685 necessario per controllare 4× MG90S braccia"],
    ["4. TIMING:", "ZERO impatto - stampa 3D bottleneck (40-60h)"],
    ["5. INVESTIMENTO:", "€79 ora vs €78+ dopo se fallisce + tempo perso"]
]

for item in why_items:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=1).font = Font(bold=True)
    ws.cell(row=recap_start, column=2).alignment = Alignment(wrap_text=True)
    recap_start += 1

# Action plan
recap_start += 2
ws.merge_cells(f'A{recap_start}:H{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "🛒 PIANO AZIONE ACQUISTO DOMANI (14 GENNAIO)"
section_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
section_cell.font = Font(bold=True, size=12, color="000000")
section_cell.alignment = Alignment(horizontal='center')

recap_start += 1
action_items = [
    ["1.", "Apri Amazon.it carrello"],
    ["2.", "RIMUOVI: Raspberry Pi Zero 2W (€23.80)"],
    ["3.", "RIMUOVI: Amazon PowerFast Micro USB (€9.99)"],
    ["4.", "AGGIUNGI: Raspberry Pi 4 Model B 4GB (cerca 'Raspberry Pi 4 4GB')"],
    ["5.", "AGGIUNGI: Alimentatore USB-C 5V 3A (cerca 'Raspberry Pi 4 alimentatore USB-C')"],
    ["6.", "AGGIUNGI: Case alluminio + dissipatore (cerca 'Raspberry Pi 4 case alluminio')"],
    ["7.", "AGGIUNGI: PCA9685 PWM Driver (cerca 'GERUI PCA9685 PWM' - €10.09)"],
    ["8.", "Verifica totale carrello = Budget originale + €79.34"],
    ["9.", "Procedi checkout"],
    ["", ""],
    ["RISULTATO:", "Hardware ottimizzato, ZERO rischi, pronto per build!"]
]

for item in action_items:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    if item[0].endswith(".") and item[0] != "":
        ws.cell(row=recap_start, column=1).font = Font(bold=True)
    if "RISULTATO" in item[0]:
        ws.cell(row=recap_start, column=1).font = Font(bold=True, color="008000")
        ws.cell(row=recap_start, column=2).font = Font(bold=True, color="008000")
    recap_start += 1

# Final confirmation
recap_start += 1
ws.merge_cells(f'A{recap_start}:H{recap_start}')
final_cell = ws.cell(row=recap_start, column=1)
final_cell.value = "✅ TRACKER PRONTO - ACQUISTO UNICO DOMANI COMPLETO - ZERO RISCHI HARDWARE"
final_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
final_cell.font = Font(bold=True, size=12)
final_cell.alignment = Alignment(horizontal='center')

# Adjust column widths
ws.column_dimensions['B'].width = 50
ws.column_dimensions['F'].width = 60

# Save
wb.save('OPENDUCK_V3_FINAL_TRACKER.xlsx')

print("✅ Tracker Excel aggiornato con RECAP COMPLETO!")
print("")
print("📋 MODIFICHE EFFETTUATE:")
print("   ✓ Marcati 2 items DA RIMUOVERE (Pi Zero 2W, Micro-USB)")
print("   ✓ Aggiunti 4 items DA AGGIUNGERE (Pi 4, USB-C, Case, PCA9685)")
print("   ✓ Creato sezione RECAP completa con:")
print("      - Step 1: Items da rimuovere (€-33.79)")
print("      - Step 2: Items da aggiungere (€+113.13)")
print("      - Riepilogo costi (+€79.34 netto)")
print("      - Motivazioni tecniche upgrade")
print("      - Piano azione step-by-step domani")
print("")
print("💰 COSTO NETTO UPGRADE: +€79.34")
print("")
print("🛒 PRONTO PER ACQUISTO UNICO DOMANI (14 GENNAIO)!")
print("   Il tracker ora contiene tutto il necessario per fare")
print("   l'acquisto in una volta sola senza dimenticare nulla.")
