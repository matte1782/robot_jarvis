import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Fix Windows console encoding
sys.stdout.reconfigure(encoding='utf-8')

# Load the workbook
wb = openpyxl.load_workbook('OPENDUCK_V3_FINAL_TRACKER.xlsx')
ws = wb.active

# Find column indices
col_map = {}
for idx, cell in enumerate(ws[1], start=1):
    if cell.value:
        header = str(cell.value).upper()
        if 'COMPONENTE' in header or 'NOME' in header or 'ITEM' in header:
            col_map['name'] = idx
        elif 'QTY' in header or 'QTÀ' in header or 'QUANTITÀ' in header:
            col_map['qty'] = idx
        elif 'PREZZO' in header or 'PRICE' in header or 'COSTO' in header:
            col_map['price'] = idx
        elif 'SOURCE' in header or 'FORNITORE' in header:
            col_map['source'] = idx
        elif 'STATUS' in header or 'STATO' in header:
            col_map['status'] = idx
        elif 'NOTE' in header or 'NOTES' in header:
            col_map['note'] = idx

# Mark items to REMOVE
items_to_remove = [
    ('Raspberry Pi Zero 2W', '⚠️ DA RIMUOVERE', 'Upgrade a Pi 4 4GB necessario (CPU insufficiente)'),
    ('micro', '⚠️ DA RIMUOVERE', 'Pi 4 usa USB-C, non micro-USB')
]

for row in ws.iter_rows(min_row=2, max_row=100):
    if not row[col_map['name']-1].value:
        continue

    item_name = str(row[col_map['name']-1].value).lower()

    for search_term, new_status, note in items_to_remove:
        if search_term.lower() in item_name:
            # Update status
            if 'status' in col_map:
                cell = row[col_map['status']-1]
                cell.value = new_status
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.font = Font(bold=True, color="000000")

            # Update note
            if 'note' in col_map:
                row[col_map['note']-1].value = note

# Find last row with data
last_row = 2
for row in range(2, 200):
    if ws.cell(row=row, column=col_map['name']).value:
        last_row = row

# Add new items BEFORE the summary section
new_items_start = last_row + 2

# Items to ADD
new_items = [
    {
        'name': '🔄 Raspberry Pi 4 Model B 4GB RAM',
        'qty': 1,
        'price': 76.60,
        'source': 'Amazon.it',
        'status': '✅ DA AGGIUNGERE',
        'note': 'Sostituisce Pi Zero 2W. CPU 3-4× più veloce, elimina bottleneck'
    },
    {
        'name': '🔄 Alimentatore USB-C 5V 3A Raspberry Pi 4',
        'qty': 1,
        'price': 13.25,
        'source': 'Amazon.it',
        'status': '✅ DA AGGIUNGERE',
        'note': 'Pi 4 richiede USB-C (non micro-USB). Necessario per alimentazione.'
    },
    {
        'name': '🔄 Case Alluminio + Dissipatore Passivo Pi 4',
        'qty': 1,
        'price': 13.19,
        'source': 'Amazon.it',
        'status': '✅ DA AGGIUNGERE',
        'note': 'Previene thermal throttling (80°C). Mantiene CPU a 65°C max.'
    },
    {
        'name': '➕ PCA9685 PWM Driver Board 16-Ch (2pcs)',
        'qty': 1,
        'price': 10.09,
        'source': 'Amazon.it - GERUI',
        'status': '✅ DA AGGIUNGERE',
        'note': 'MANCANTE dal BOM! Necessario per controllare 4× MG90S (braccia). I2C addr 0x40'
    }
]

# Add new items
for i, item in enumerate(new_items, start=1):
    row_num = new_items_start + i - 1

    if 'name' in col_map:
        cell = ws.cell(row=row_num, column=col_map['name'])
        cell.value = item['name']
        cell.font = Font(bold=True)

    if 'qty' in col_map:
        ws.cell(row=row_num, column=col_map['qty']).value = item['qty']

    if 'price' in col_map:
        ws.cell(row=row_num, column=col_map['price']).value = item['price']

    if 'source' in col_map:
        ws.cell(row=row_num, column=col_map['source']).value = item['source']

    if 'status' in col_map:
        cell = ws.cell(row=row_num, column=col_map['status'])
        cell.value = item['status']
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.font = Font(bold=True, color="000000")

    if 'note' in col_map:
        cell = ws.cell(row=row_num, column=col_map['note'])
        cell.value = item['note']
        cell.alignment = Alignment(wrap_text=True)

# Create comprehensive RECAP section
recap_start = new_items_start + len(new_items) + 2

# Header style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=14, color="FFFFFF")
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=11)

# Main header
ws.merge_cells(f'A{recap_start}:F{recap_start}')
header_cell = ws.cell(row=recap_start, column=1)
header_cell.value = "📋 RECAP COMPLETO MODIFICHE - ACQUISTO UNICO DOMANI"
header_cell.fill = header_fill
header_cell.font = header_font
header_cell.alignment = Alignment(horizontal='center', vertical='center')

# Section 1: Items to REMOVE
recap_start += 2
ws.merge_cells(f'A{recap_start}:F{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "❌ STEP 1: RIMUOVI DAL CARRELLO"
section_cell.fill = section_fill
section_cell.font = section_font

recap_start += 1
remove_items = [
    ["Raspberry Pi Zero 2W", "€23.80", "CPU insufficiente (47-83% carico), upgrade necessario"],
    ["Cavo Micro-USB", "€9.99", "Pi 4 usa USB-C, non compatibile"]
]

for item in remove_items:
    ws.cell(row=recap_start, column=1).value = f"⚠️ {item[0]}"
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=3).value = item[2]
    ws.cell(row=recap_start, column=1).font = Font(bold=True, color="C00000")
    recap_start += 1

# Section 2: Items to ADD
recap_start += 1
ws.merge_cells(f'A{recap_start}:F{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "✅ STEP 2: AGGIUNGI AL CARRELLO"
section_cell.fill = section_fill
section_cell.font = section_font

recap_start += 1
add_items = [
    ["Raspberry Pi 4 Model B 4GB", "€76.60", "3-4× più veloce, elimina bottleneck CPU"],
    ["Alimentatore USB-C 5V 3A", "€13.25", "Alimentazione corretta per Pi 4"],
    ["Case Alluminio + Dissipatore", "€13.19", "Previene thermal throttling"],
    ["PCA9685 PWM Driver (2pcs)", "€10.09", "Controller I2C per servos braccia MG90S"]
]

for item in add_items:
    ws.cell(row=recap_start, column=1).value = f"✅ {item[0]}"
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=3).value = item[2]
    ws.cell(row=recap_start, column=1).font = Font(bold=True, color="008000")
    recap_start += 1

# Section 3: Cost Summary
recap_start += 1
ws.merge_cells(f'A{recap_start}:F{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "💰 RIEPILOGO COSTI"
section_cell.fill = section_fill
section_cell.font = section_font

recap_start += 1
cost_summary = [
    ["Items RIMOSSI:", "-€33.79", "(Pi Zero 2W + Micro-USB)"],
    ["Items AGGIUNTI:", "+€113.13", "(Pi 4 + USB-C + Case + PCA9685)"],
    ["", "", ""],
    ["COSTO NETTO UPGRADE:", "+€79.34", "Investimento per eliminare rischi"]
]

for item in cost_summary:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=3).value = item[2]
    if "NETTO" in item[0]:
        ws.cell(row=recap_start, column=1).font = Font(bold=True, size=12)
        ws.cell(row=recap_start, column=2).font = Font(bold=True, size=12, color="C00000")
    recap_start += 1

# Section 4: Why this upgrade
recap_start += 1
ws.merge_cells(f'A{recap_start}:F{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "🔍 PERCHÉ QUESTO UPGRADE?"
section_cell.fill = section_fill
section_cell.font = section_font

recap_start += 1
why_items = [
    ["BOTTLENECK TROVATO:", "Pi Zero 2W CPU carico 47-83% single core"],
    ["RISCHIO POTENZA:", "5V rail marginal con braccia (1.3A vs 3A max)"],
    ["SOLUZIONE:", "Pi 4 ha 3-4× CPU, 8× RAM, maggiore headroom"],
    ["COMPONENTE MANCANTE:", "PCA9685 necessario per controllare MG90S braccia"],
    ["TIMING:", "NO impatto - stampa 3D è bottleneck (40-60h)"]
]

for item in why_items:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    ws.cell(row=recap_start, column=1).font = Font(bold=True)
    recap_start += 1

# Section 5: Shopping action plan
recap_start += 1
ws.merge_cells(f'A{recap_start}:F{recap_start}')
section_cell = ws.cell(row=recap_start, column=1)
section_cell.value = "🛒 PIANO AZIONE DOMANI"
section_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
section_cell.font = Font(bold=True, size=12)

recap_start += 1
action_plan = [
    ["1.", "Vai su Amazon.it carrello"],
    ["2.", "RIMUOVI: Pi Zero 2W + Micro-USB cable"],
    ["3.", "AGGIUNGI: Pi 4 4GB + USB-C alimentatore + Case alluminio"],
    ["4.", "AGGIUNGI: PCA9685 PWM Driver (GERUI 2pcs €10.09)"],
    ["5.", "Verifica totale carrello e checkout"],
    ["", ""],
    ["LINK DIRETTI:", ""],
    ["Pi 4 4GB:", "Cerca 'Raspberry Pi 4 Model B 4GB'"],
    ["USB-C Power:", "Cerca 'Raspberry Pi 4 alimentatore USB-C 5V 3A'"],
    ["Case:", "Cerca 'Raspberry Pi 4 case alluminio dissipatore'"],
    ["PCA9685:", "Cerca 'GERUI PCA9685' o 'PCA9685 PWM driver'"]
]

for item in action_plan:
    ws.cell(row=recap_start, column=1).value = item[0]
    ws.cell(row=recap_start, column=2).value = item[1]
    if item[0].endswith("."):
        ws.cell(row=recap_start, column=1).font = Font(bold=True)
    recap_start += 1

# Final note
recap_start += 1
ws.merge_cells(f'A{recap_start}:F{recap_start}')
note_cell = ws.cell(row=recap_start, column=1)
note_cell.value = "✅ PRONTO PER ACQUISTO UNICO DOMANI - Tutti i rischi hardware eliminati"
note_cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
note_cell.font = Font(bold=True, size=11)
note_cell.alignment = Alignment(horizontal='center')

# Adjust column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 30

# Save
wb.save('OPENDUCK_V3_FINAL_TRACKER.xlsx')

print("✅ Tracker aggiornato con RECAP COMPLETO!")
print("")
print("📋 MODIFICHE EFFETTUATE:")
print("   1. Marcati items DA RIMUOVERE (Pi Zero 2W, Micro-USB)")
print("   2. Aggiunti 4 items DA AGGIUNGERE (Pi 4, USB-C, Case, PCA9685)")
print("   3. Creato RECAP sezione completa con:")
print("      - Step 1: Items da rimuovere")
print("      - Step 2: Items da aggiungere")
print("      - Riepilogo costi (+€79.34 netto)")
print("      - Motivazioni upgrade")
print("      - Piano azione domani")
print("")
print("💰 COSTO UPGRADE NETTO: +€79.34")
print("   - Rimossi: -€33.79")
print("   - Aggiunti: +€113.13")
print("")
print("🛒 PRONTO PER ACQUISTO UNICO DOMANI!")
